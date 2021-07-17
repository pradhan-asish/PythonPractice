import openpyxl,os
wb = openpyxl.Workbook()
print(wb)
print(wb.sheetnames)
sheet1 = wb['Sheet']
print(sheet1['A1'].value)

sheet1['A1'].value = 10
sheet1['A2'].value = 12
sheet1['A3'].value = 13
sheet1['A4'].value = 14

wb.save('Example1.xlsx')

wb1 = openpyxl.load_workbook('Example1.xlsx')
sheet1 = wb['Sheet']

for i in range(1,5):
    print(sheet1.cell(row = i , column = 1).value)

print(wb1.sheetnames)

wb1.create_sheet('New Sheet 1')

print(wb1.sheetnames)

sheet2 = wb1['New Sheet 1']
sheet2.title = 'New Sheet 2'
    
print(wb1.sheetnames)
